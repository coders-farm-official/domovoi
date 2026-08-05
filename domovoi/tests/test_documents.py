"""Tests for the homegrown Documents suite (web/backend/api/documents.py).

Three concerns:
  * Containment — every path resolves inside documents_dir; traversal /
    absolute paths are rejected (the music.py:349-367 pattern).
  * The homegrown editors' data paths — text read/write, the sheet grid
    round-trip (.csv values + .xlsx formulas via openpyxl), and category
    routing for the list view.
  * Exports — markdown → .docx (python-docx) and sheet → csv/xlsx.
"""

from __future__ import annotations

import io
import zipfile

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient

from domovoi.config import settings
from web.backend.api import documents as docs
from web.backend.main import app


@pytest.fixture
def docs_dir(monkeypatch, tmp_path):
    monkeypatch.setattr(settings, "documents_dir", str(tmp_path))
    return tmp_path


def _client():
    return TestClient(app)


# ─── Containment ────────────────────────────────────────────────────
def test_safe_target_accepts_plain_name(docs_dir):
    target = docs._safe_target("report.md")
    assert target == (docs_dir / "report.md").resolve()


def test_safe_target_rejects_parent_traversal(docs_dir):
    with pytest.raises(HTTPException) as ei:
        docs._safe_target("../../etc/passwd")
    assert ei.value.status_code == 400


def test_safe_target_rejects_absolute_escape(docs_dir):
    outside = docs_dir.parent / "somewhere_else" / "x.md"
    with pytest.raises(HTTPException) as ei:
        docs._safe_target(str(outside))
    assert ei.value.status_code == 400


def test_safe_target_rejects_empty(docs_dir):
    with pytest.raises(HTTPException):
        docs._safe_target("")


def test_rel_of_roundtrips(docs_dir):
    t = docs._safe_target("a.md")
    assert docs._rel_of(t) == "a.md"


# ─── Category routing ───────────────────────────────────────────────
def test_doc_category_routing():
    assert docs._doc_category(".md") == "doc"
    assert docs._doc_category(".xlsx") == "sheet"
    assert docs._doc_category(".csv") == "sheet"
    assert docs._doc_category(".docx") == "download"
    assert docs._doc_category(".xls") == "download"
    assert docs._doc_category(".ods") == "download"
    assert docs._doc_category(".excalidraw") == "drawing"
    assert docs._doc_category(".pdf") == "newtab"
    assert docs._doc_category(".png") == "newtab"
    assert docs._doc_category(".txt") == "text"
    assert docs._doc_category(".weird") == "text"


# ─── Create / list ──────────────────────────────────────────────────
def test_create_doc_defaults_to_markdown(docs_dir):
    with _client() as c:
        r = c.post("/api/documents/create", json={"name": "notes", "kind": "doc"})
    assert r.status_code == 200
    body = r.json()
    assert body["rel_path"] == "notes.md"
    assert body["category"] == "doc"
    assert (docs_dir / "notes.md").is_file()


def test_create_sheet_is_valid_xlsx(docs_dir):
    with _client() as c:
        r = c.post("/api/documents/create", json={"name": "budget", "kind": "sheet"})
    assert r.json()["rel_path"] == "budget.xlsx"
    from openpyxl import load_workbook

    wb = load_workbook(docs_dir / "budget.xlsx")
    assert wb.worksheets  # opens cleanly


def test_list_tags_categories(docs_dir):
    (docs_dir / "a.md").write_text("# hi", encoding="utf-8")
    (docs_dir / "b.csv").write_text("1,2", encoding="utf-8")
    (docs_dir / "legacy.docx").write_bytes(b"pk")
    with _client() as c:
        rows = {r["rel_path"]: r for r in c.get("/api/documents/").json()}
    assert rows["a.md"]["category"] == "doc"
    assert rows["b.csv"]["category"] == "sheet"
    assert rows["legacy.docx"]["category"] == "download"


# ─── Sheet grid round-trip ──────────────────────────────────────────
def test_sheet_csv_roundtrip(docs_dir):
    (docs_dir / "t.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    with _client() as c:
        got = c.get("/api/documents/sheet/t.csv").json()
        assert [[cell["v"] for cell in row] for row in got["rows"]] == [["a", "b"], ["1", "2"]]

        r = c.put("/api/documents/sheet/t.csv", json={"rows": [
            [{"v": "x"}, {"v": "9"}],
            [{"f": "=SUM(B1)"}, None],
        ]})
        assert r.status_code == 200
    text_ = (docs_dir / "t.csv").read_text(encoding="utf-8")
    assert text_.splitlines() == ["x,9", "=SUM(B1),"]


def test_sheet_xlsx_roundtrip_keeps_formulas(docs_dir):
    with _client() as c:
        c.post("/api/documents/create", json={"name": "calc", "kind": "sheet"})
        r = c.put("/api/documents/sheet/calc.xlsx", json={"rows": [
            [{"v": "3"}, {"v": "4"}, {"f": "=SUM(A1:B1)"}],
        ]})
        assert r.status_code == 200
        got = c.get("/api/documents/sheet/calc.xlsx").json()
    row = got["rows"][0]
    assert row[0]["v"] == "3" and row[1]["v"] == "4"
    assert row[2]["f"] == "=SUM(A1:B1)"


def test_sheet_unsupported_type_415(docs_dir):
    (docs_dir / "old.xls").write_bytes(b"\xd0\xcf")
    with _client() as c:
        assert c.get("/api/documents/sheet/old.xls").status_code == 415


# ─── Exports ────────────────────────────────────────────────────────
def test_export_markdown_to_docx(docs_dir):
    (docs_dir / "doc.md").write_text(
        "# Title\n\nSome **bold** and *italic* text.\n\n- one\n- two\n\n"
        "```\ncode here\n```\n",
        encoding="utf-8",
    )
    with _client() as c:
        r = c.get("/api/documents/export/doc/doc.md", params={"fmt": "docx"})
    assert r.status_code == 200
    assert "wordprocessingml" in r.headers["content-type"]
    assert 'filename="doc.docx"' in r.headers["content-disposition"]
    # It's a real docx: python-docx can open it and the heading survived.
    from docx import Document

    d = Document(io.BytesIO(r.content))
    texts = [p.text for p in d.paragraphs]
    assert "Title" in texts
    assert any("bold" in t for t in texts)


def test_export_sheet_csv_and_xlsx(docs_dir):
    (docs_dir / "t.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    with _client() as c:
        csv_r = c.get("/api/documents/export/sheet/t.csv", params={"fmt": "csv"})
        xlsx_r = c.get("/api/documents/export/sheet/t.csv", params={"fmt": "xlsx"})
    assert csv_r.status_code == 200
    assert csv_r.text.splitlines() == ["a,b", "1,2"]
    assert xlsx_r.status_code == 200
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_r.content))
    ws = wb.worksheets[0]
    assert ws["A1"].value == "a" and ws["B2"].value == 2


# ─── Text editor + raw + zip (retained surface) ─────────────────────
def test_text_roundtrip_and_binary_guard(docs_dir):
    (docs_dir / "n.txt").write_text("hello", encoding="utf-8")
    (docs_dir / "blob.bin").write_bytes(b"\x00\xff\x00\xff")
    with _client() as c:
        ok = c.get("/api/documents/text/n.txt").json()
        assert ok["text"] == "hello"
        w = c.put("/api/documents/text/n.txt", json={"text": "changed"})
        assert w.status_code == 200
        assert (docs_dir / "n.txt").read_text(encoding="utf-8") == "changed"
        binary = c.get("/api/documents/text/blob.bin")
        assert binary.status_code == 415
        assert binary.json()["reason"] == "binary"


def test_zip_and_delete(docs_dir):
    (docs_dir / "a.md").write_text("a", encoding="utf-8")
    (docs_dir / "b.md").write_text("b", encoding="utf-8")
    with _client() as c:
        z = c.post("/api/documents/download-zip", json={"rel_paths": ["a.md", "b.md"]})
        assert z.status_code == 200
        names = zipfile.ZipFile(io.BytesIO(z.content)).namelist()
        assert sorted(names) == ["a.md", "b.md"]

        d = c.post("/api/documents/delete", json={"rel_paths": ["a.md", "missing.md"]})
        body = d.json()
        assert body["deleted"] == ["a.md"] and body["failed"] == ["missing.md"]
        assert not (docs_dir / "a.md").exists()
